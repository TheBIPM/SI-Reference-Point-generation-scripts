import openpyxl
import os
from rdflib import Graph, URIRef, Namespace

from settings import APIPATH

# load existing decisions.ttl
g = Graph()
g.parse(os.path.join("CUQ",
                     "_docs",
                     "decisions_without_cipm.ttl"), format="ttl")

# add required namespaces
CIPM = Namespace("http://si-digital-framework.org/bodies/CIPM#")
CCTF = Namespace("http://si-digital-framework.org/bodies/CCTF#")
DEC = Namespace("http://si-digital-framework.org/Decisions#")
g.bind("cipm", CIPM)
g.bind("cctf", CCTF)
g.bind("dec", DEC)

# load missing entries from excel file
missing_entries_file = openpyxl.load_workbook(os.path.join(
    "CUQ", "_docs", "missing_definitions_cipm.xlsx"))
ws = missing_entries_file["Feuil1"]
dec_ids = ws["A"]
res_ids = ws["F"]

ns_dict = dict(g.namespaces())

for sub, obj in zip(dec_ids[1:], res_ids[1:]):
    #print(f"{sub.value :13s} -->   {obj.value}")

    # split the object string
    obj_ns, obj_ref = obj.value.split(":")
    OBJ_NAMESPACE = Namespace(ns_dict[obj_ns])

    g.add(
        (
            URIRef(DEC.term(sub.value)),
            URIRef(DEC.correspondingResolution),
            URIRef(OBJ_NAMESPACE.term(obj_ref)),
        )
    )

# output
g.serialize(format="turtle", destination=APIPATH + "decisions.ttl")
