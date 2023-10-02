#
# Quantities ABox
#
#
# the module imports the Python script 'settings.py' located in a directory above this directory
# Should this lead to an error when executing the present script (... not found ...) you might need
# to add the location to the PYTHONPATH by typing (in the TERMINAL where you execute the Python script)
# export PYTHONPATH=(path where the package is located) (e.g. /Users/gregordudle/Development/Semantic-SI)
#

from rdflib import URIRef, Literal, Graph
from rdflib.namespace import DCTERMS, RDF, RDFS, SKOS, OWL, XSD
from CUQ_TBox import SiElements
from datetime import date
from settings import *
import openpyxl


PDF = SiElements()
g = Graph()

# copy over all namespaces from PDF.g to g
for key, val in PDF.g.namespaces():
    g.bind(key, val) 

# Annotations to the ontology (name, Version number)
g.add((URIRef(PDF.namespace_quantities), RDF.type, OWL.Ontology))
g.add((URIRef(PDF.namespace_quantities), SKOS.prefLabel, Literal("SI Reference Point - Quantities", datatype=XSD.string)))
g.add((URIRef(PDF.namespace_quantities), RDFS.comment, Literal("Ontology, part of the SI reference point, covering quantities",
                                                 datatype=XSD.string)))
g.add((URIRef(PDF.namespace_quantities), DCTERMS.created, Literal(str(date.today()), datatype=XSD.date)))

# crawl through the rows of the XLS file

# worksheet containing the basic information
qty_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'quantities.xlsx')
sheet = qty_wb_obj.active

# crawl through the rows of the XLS file
for row in range(2, sheet.max_row):
    identifier = sheet.cell(row, column=1).value
    label_en = sheet.cell(row, column=2).value
    label_fr = sheet.cell(row, column=3).value
    
    element = PDF.set_quantity_uri(identifier)
    g.add((element, RDF.type, PDF.QuantityKind))
    g.add((element, SKOS.prefLabel, Literal(label_en, lang="en")))
    g.add((element, SKOS.prefLabel, Literal(label_fr, lang="fr")))
    g.add((element, SKOS.altLabel, Literal(identifier, datatype=XSD.string)))
    if sheet.cell(row, column=4).value is not None:
        unit = sheet.cell(row, column=4).value
        g.add((element, PDF.hasUnit, PDF.set_unit_uri(unit)))
    if sheet.cell(row, column=5).value is not None:
        sameas = sheet.cell(row, column=5).value
        g.add((element, OWL.sameAs, PDF.set_uri(sameas)))
    if sheet.cell(row, column=6).value is not None:
        sameas = sheet.cell(row, column=6).value
        g.add((element, OWL.sameAs, PDF.set_uri(sameas)))


g.serialize(format='turtle', destination=APIPATH + 'quantities.ttl')
