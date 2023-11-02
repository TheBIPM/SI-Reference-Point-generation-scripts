"""
Units ABox
The module imports the Python script 'settings.py' located in a directory above this directory
should this lead to an error when executing the present script (... not found ...) you might need
to add the location to the PYTHONPATH by typing (in the TERMINAL where you execute the Python script)
export PYTHONPATH=(path where the package is located) (e.g. /Users/gregordudle/Development/Semantic-SI)
"""

from rdflib import *
from CUQ_TBox import SiElements
from datetime import date
from settings import *
import openpyxl
import re

# import CUQ TBox
PDF = SiElements()
g = Graph()

# copy over all namespaces from PDF.g to g
for key, val in PDF.g.namespaces():
    g.bind(key, val)

BASESTR = str(PDF.BASE_PATH)

# Annotations to the ontology (name, Version number)
g.add((URIRef(PDF.namespace_units), RDF.type, OWL.Ontology))
g.add((URIRef(PDF.namespace_units), SKOS.prefLabel,
       Literal("SI Reference Point - Units and Prefixes", datatype=XSD.string)))
g.add((URIRef(PDF.namespace_units), RDFS.comment,
       Literal("Ontology, part of the SI Reference Point, covering measurement units "
               "(SI base units and SI units with special names) and prefixes.")))
g.add((URIRef(PDF.namespace_units), DCTERMS.created, Literal(str(date.today()), datatype=XSD.date)))

# 1) open XLS files with information
units_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'Units_Prefixes.xlsx')
notes_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'Notes.xlsx')

# 2) create dictionary with the note symbol codes (@xxx@) in text
notessym = notes_wb_obj["symbols"]
syms, symtypes, formats = {}, {}, {'html': 6, 'latex': 7, 'json': 8, 'text': 9}
for symrow in range(2, notessym.max_row + 1):
    symtypes.update({notessym.cell(symrow, column=5).value: notessym.cell(symrow, column=4).value})
    for ttype in ['latex']:
        if ttype not in syms.keys():
            syms.update({ttype: {}})
        code = notessym.cell(symrow, column=5).value
        text = notessym.cell(symrow, column=formats[ttype]).value
        syms[ttype].update({code: text})


# 3) Note function to change out symbol references in notes (@xxx@), uses syms and symtypes vars defined above
def formattxt(txt, fmt='html', lvl=0):
    """
    formats a text string with symbols replaced in any of the following formats
    html, latex, json, text
    """
    if fmt not in ['html', 'latex', 'json', 'text']:
        return txt

    matches = re.findall(f"@(.*?)@", txt)
    # order matches by length of substituting string (smallest first) so that replacement of substrings
    # at the top level does not result in multiple replacements at top level (resulting in $$)
    tmatches = {}
    for match in matches:
        strlen = len(syms[fmt][match])
        tmatches.update({match: strlen})
    omatches = dict(sorted(tmatches.items(), key=lambda item: item[1]))
    if omatches:
        for grp in omatches:
            if lvl == 0 and fmt == 'latex':
                # add latex delimiters at the top level
                if symtypes[grp] == 'symbol':
                    txt = txt.replace('@' + grp + '@', "$" + syms[fmt][grp] + "$")
                elif symtypes[grp] == 'equation':
                    txt = txt.replace('@' + grp + '@', syms[fmt][grp])
            else:
                txt = txt.replace('@' + grp + '@', syms[fmt][grp])
        txt = txt.replace('\n', ' ')
        # check for text still containing symbols based on symbols being parts of other symbols
        if '@' in txt:
            lvl = lvl + 1
            txt = formattxt(txt, fmt, lvl)
    return txt


# 4) Base unit definitions

# 4.1) Define the BaseUnit (to which one can subsequently attach several definitions)
sheet = units_wb_obj["DefCollectors"]
for row in range(2, sheet.max_row + 1):
    uri_text = sheet.cell(row=row, column=1).value
    prefLabel_fr = sheet.cell(row, column=2).value
    prefLabel_en = sheet.cell(row, column=3).value
    UnitOfQtyKind = sheet.cell(row, column=4).value
    symbol = sheet.cell(row, column=5).value

    if uri_text is not None:
        element = PDF.set_unit_uri(uri_text)
        g.add((element, RDF.type, PDF.SIBaseUnit))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_fr, lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_en, lang='en')))
        g.add((element, PDF.hasUnitTypeAsString, Literal('SI base unit', lang='en')))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Unité SI de base', lang='fr')))
        g.add((element, PDF.isUnitOfQtyKind, PDF.set_quantity_uri(UnitOfQtyKind)))
        g.add((element, PDF.hasSymbol, Literal(symbol, datatype=XSD.string)))

        for kolonne in range(6, sheet.max_column + 1):
            curr_def = sheet.cell(row, column=kolonne).value
            next_def = sheet.cell(row, column=kolonne + 1).value
            if curr_def is not None:
                g.add((element, PDF.hasDefinition, PDF.set_uri(curr_def)))
                if next_def is not None:
                    g.add((PDF.set_uri(curr_def), PDF.hasNextDefinition, PDF.set_uri(next_def)))

# 4.2 Declare all definitions
# uri_text values are a concatenation of the lowercase unit name and the year of the definition, e.g., ampere2018
basedefs = units_wb_obj["BaseUnitsDefs"]
notesen = notes_wb_obj["en"]
notesfr = notes_wb_obj["fr"]

for row in range(2, basedefs.max_row + 1):
    # load data
    uri_text = basedefs.cell(row, column=1).value
    prefLabel_fr = basedefs.cell(row, column=2).value
    prefLabel_en = basedefs.cell(row, column=3).value
    StartValidity = basedefs.cell(row, column=4).value
    EndValidity = basedefs.cell(row, column=5).value
    DefiningText_fr = basedefs.cell(row, column=6).value
    DefiningText_en = basedefs.cell(row, column=7).value
    DefiningResolution = basedefs.cell(row, column=8).value
    DefiningEquation = basedefs.cell(row, column=9).value
    DefiningConstant = basedefs.cell(row, column=10).value
    UnitOfQtyKind = basedefs.cell(row, column=11).value
    Status = basedefs.cell(row, column=12).value

    # add data
    if uri_text is not None:
        element = PDF.set_uri(uri_text)
        g.add((element, RDF.type, PDF.Definition))
        g.add((element, PDF.hasUnitTypeAsString, Literal('SI base unit', lang='en')))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Unité SI de base', lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_fr, lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_en, lang='en')))
        g.add((element, PDF.hasStartValidity, Literal(StartValidity, datatype=XSD.date)))
        if EndValidity is not None:
            g.add((element, PDF.hasEndValidity, Literal(EndValidity, datatype=XSD.date)))
        if DefiningText_fr is not None:
            # change any symbols to latex
            DefiningText_fr = formattxt(DefiningText_fr, 'latex')
            g.add((element, PDF.hasDefiningText, Literal(DefiningText_fr, lang='fr')))
        if DefiningText_en is not None:
            # change any symbols to latex
            DefiningText_en = formattxt(DefiningText_en, 'latex')
            g.add((element, PDF.hasDefiningText, Literal(DefiningText_en, lang='en')))
        g.add((element, PDF.hasDefiningResolution, URIRef(PDF.set_cgpm_uri(DefiningResolution))))
        if DefiningEquation is not None:
            if "@" in DefiningEquation:
                DefiningEquation = formattxt(DefiningEquation, 'latex', 1)
            g.add((element, PDF.hasDefiningEquation, Literal(DefiningEquation, datatype=XSD.string)))
        if DefiningConstant is not None:
            g.add((element, PDF.hasDefiningConstant, PDF.set_constant_uri(DefiningConstant)))
        if Status is not None:
            g.add((element, PDF.hasStatus, Literal(Status, datatype=XSD.string)))

        # notes
        # get all the notes for a definition
        temp = {}
        for noterow in range(2, notesen.max_row + 1):
            if notesen.cell(noterow, column=2).value == uri_text and notesen.cell(noterow, column=3).value is not None:
                temp.update({notesen.cell(noterow, column=3).value: notesen.cell(noterow, column=4).value})
        notes = dict(sorted(temp.items()))

        temp_fr = {}
        for noterow in range(2, notesfr.max_row + 1):
            if notesfr.cell(noterow, column=2).value == uri_text and notesfr.cell(noterow, column=3).value is not None:
                temp_fr.update({notesfr.cell(noterow, column=3).value: notesfr.cell(noterow, column=4).value})
        notes_fr = dict(sorted(temp_fr.items()))

        # add all the notes for a definition
        if notes:
            for nidx, note in notes.items():
                if "@" in note:
                    note = formattxt(note, 'latex')

                note_uri = uri_text + "note" + str(nidx)
                notenode = PDF.set_uri(note_uri)
                g.add((element, PDF.hasDefinitionNote, notenode))
                g.add((notenode, RDF.type, PDF.DefinitionNote))
                g.add((notenode, PDF.hasNoteIndex, Literal(nidx)))
                g.add((notenode, PDF.hasNoteText, Literal(note, lang='en')))

        if notes_fr:
            for nidx, note_fr in notes_fr.items():
                if "@" in note_fr:
                    note_fr = formattxt(note_fr, 'latex')

                note_uri = uri_text + "note" + str(nidx)
                notenode = PDF.set_uri(note_uri)
                g.add((notenode, PDF.hasNoteText, Literal(note_fr, lang='fr')))

# 5 SI Units Special Names
sheet = units_wb_obj["SIUnitsSpecialNames"]

for row in range(2, sheet.max_row + 1):
    # load data
    uri_text = sheet.cell(row, column=1).value
    prefLabel_fr = sheet.cell(row, column=2).value
    prefLabel_en = sheet.cell(row, column=3).value
    symbol = sheet.cell(row, column=4).value
    defres = sheet.cell(row, column=5).value
    UnitOfQtyKind = sheet.cell(row, column=6).value
    othersi = sheet.cell(row, column=7).value
    inbasesi = sheet.cell(row, column=8).value
    equation = sheet.cell(row, column=9).value

    # add data
    if uri_text is not None:
        if "@" in symbol:
            symbol = formattxt(symbol, 'latex')

        element = PDF.set_unit_uri(uri_text)
        g.add((element, RDF.type, PDF.SISpecialNamedUnit))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Named SI derived unit', lang='en')))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Unité SI dérivée ayant un nom spécial', lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_fr, lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_en, lang='en')))
        g.add((element, PDF.hasSymbol, Literal(symbol, datatype=XSD.string)))
        g.add((element, PDF.isUnitOfQtyKind, PDF.set_quantity_uri(UnitOfQtyKind)))
        g.add((element, PDF.hasDefiningResolution, URIRef(PDF.set_cgpm_uri(defres))))
        if othersi:
            if "@" in othersi:
                othersi = formattxt(othersi, 'latex', 1)
            g.add((element, PDF.inOtherSIUnits, Literal(othersi, datatype=XSD.string)))
        if inbasesi:
            if "@" in inbasesi:
                inbasesi = formattxt(inbasesi, 'latex', 1)
            g.add((element, PDF.inBaseSIUnits, Literal(inbasesi, datatype=XSD.string)))
        if equation:
            if "@" in equation:
                equation = formattxt(equation, 'latex', 1)
            g.add((element, PDF.hasDefiningEquation, Literal(equation, datatype=XSD.string)))

# 6) non SI units
sheet = units_wb_obj["NonSIUnits"]

for row in range(7, sheet.max_row + 1):
    # load data
    uri_text = sheet.cell(row, column=1).value
    prefLabel_fr = sheet.cell(row, column=2).value
    prefLabel_en = sheet.cell(row, column=3).value
    symbol = sheet.cell(row, column=4).value
    unitOfQtyKind = sheet.cell(row, column=5).value
    conversionFactor = sheet.cell(row, column=6).value
    conversionUnit = sheet.cell(row, column=7).value

    # add data
    if uri_text is not None:
        if "@" in symbol:
            symbol = formattxt(symbol, 'latex')

        element = PDF.set_unit_uri(uri_text)
        g.add((element, RDF.type, PDF.nonSIUnit))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Non-SI unit accepted for use with the SI', lang='en')))
        g.add((element, PDF.hasUnitTypeAsString, Literal('Unité en dehors du SI '
                                                         'dont l\'usage est accepté avec le SI', lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_fr, lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_en, lang='en')))
        g.add((element, PDF.hasSymbol, Literal(symbol, datatype=XSD.string)))
        g.add((element, PDF.isUnitOfQtyKind, PDF.set_quantity_uri(unitOfQtyKind)))
        g.add((element, PDF.hasConversionFactor, Literal(conversionFactor, datatype=XSD.double)))
        g.add((element, PDF.hasConversionUnit, PDF.set_unit_uri(conversionUnit)))

# 7) serialization
g.serialize(format='ttl', destination=APIPATH + 'units.ttl')
