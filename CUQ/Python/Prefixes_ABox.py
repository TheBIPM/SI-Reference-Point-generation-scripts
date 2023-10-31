#
# Prefixes ABox
#
#
# the module imports the Python script 'settings.py' located in a directory above this directory
# Should this lead to an error when executing the present script (... not found ...) you might need
# to add the location to the PYTHONPATH by typing (in the TERMINAL where you execute the Python script)
# export PYTHONPATH=(path where the package is located) (e.g. /Users/gregordudle/Development/Semantic-SI)
#

from rdflib import *
from CUQ_TBox import SiElements
from datetime import date
from settings import *
import openpyxl
import re

# import CUQ TBox
PDF = SiElements()
g = Graph()

# copy over all namespaces from PDF.g to g
for key, val in PDF.g.namespaces():
    g.bind(key, val) 

# Annotations to the ontology (name, Version number)
g.add((URIRef(PDF.namespace_prefixes), RDF.type, OWL.Ontology))
g.add((URIRef(PDF.namespace_prefixes), SKOS.prefLabel, Literal("SI Reference Point - Prefixes", datatype=XSD.string)))
g.add((URIRef(PDF.namespace_prefixes), RDFS.comment, Literal("Ontology, part of the SI Reference Point, covering "
                                                             "prefixes for the SI measurement units.")))
g.add((URIRef(PDF.namespace_prefixes), DCTERMS.created, Literal(str(date.today()), datatype=XSD.date)))

# 1) open XLS files with information
units_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'Units_Prefixes.xlsx')
notes_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'Notes.xlsx')

# 2) create dictionary with the note symbol codes (@xxx@) in text
notessym = notes_wb_obj["symbols"]
syms, symtypes, formats = {}, {}, {'html': 6, 'latex': 7, 'json': 8, 'text': 9}
for symrow in range(2, notessym.max_row + 1):
    symtypes.update({notessym.cell(symrow, column=5).value: notessym.cell(symrow, column=4).value})
    for ttype in ['latex']:
        if ttype not in syms.keys():
            syms.update({ttype: {}})
        code = notessym.cell(symrow, column=5).value
        text = notessym.cell(symrow, column=formats[ttype]).value
        syms[ttype].update({code: text})


# 3) Note function to change out symbol references in notes (@xxx@), uses syms and symtypes vars defined above
def formattxt(txt, fmt='html', lvl=0):
    """
    formats a text string with symbols replaced in any of the following formats
    html, latex, json, text
    """
    if fmt not in ['html', 'latex', 'json', 'text']:
        return txt

    matches = re.findall(f"@(.*?)@", txt)
    # order matches by length of substituting string (smallest first) so that replacement of substrings
    # at the top level does not result in multiple replacements at top level (resulting in $$)
    tmatches = {}
    for match in matches:
        strlen = len(syms[fmt][match])
        tmatches.update({match: strlen})
    omatches = dict(sorted(tmatches.items(), key=lambda item: item[1]))
    if omatches:
        for grp in omatches:
            if lvl == 0 and fmt == 'latex':
                # add latex delimiters at the top level
                if symtypes[grp] == 'symbol':
                    txt = txt.replace('@' + grp + '@', "$" + syms[fmt][grp] + "$")
                elif symtypes[grp] == 'equation':
                    txt = txt.replace('@' + grp + '@', "$$" + syms[fmt][grp] + "$$")
            else:
                txt = txt.replace('@' + grp + '@',  syms[fmt][grp])
        txt = txt.replace('\n', ' ')
        # check for text still containing symbols based on symbols being parts of other symbols
        if '@' in text:
            lvl = lvl+1
            txt = formattxt(txt, fmt, lvl)
    return txt


# 4) Create prefixes
sheet = units_wb_obj["Prefixes"]

for row in range(2, sheet.max_row + 1):
    uri_text = sheet.cell(row, column=1).value
    prefLabel_en = sheet.cell(row, column=2).value
    prefLabel_fr = sheet.cell(row, column=3).value
    scalingFactor = sheet.cell(row, column=4).value
    symbol = sheet.cell(row, column=5).value
    defres = sheet.cell(row, column=6).value
    datatype = sheet.cell(row, column=7).value

    if uri_text is not None:
        element = PDF.set_prefix_uri(uri_text)
        g.add((element, RDF.type, PDF.SIPrefix))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_fr, lang='fr')))
        g.add((element, SKOS.prefLabel, Literal(prefLabel_en, lang='en')))
        if datatype == "integer":
            g.add((element, PDF.hasScalingFactor, Literal(scalingFactor, datatype=XSD.integer)))
            g.add((element, PDF.hasDatatype, XSD.integer))
        elif datatype == "decimal":
            g.add((element, PDF.hasScalingFactor, Literal(scalingFactor, datatype=XSD.decimal)))
            g.add((element, PDF.hasDatatype, XSD.decimal))
        elif datatype == "float":
            g.add((element, PDF.hasScalingFactor, Literal(scalingFactor, datatype=XSD.float)))
            g.add((element, PDF.hasDatatype, XSD.float))
        g.add((element, PDF.hasSymbol, Literal(symbol, datatype=XSD.string)))
        g.add((element, PDF.hasDefiningResolution, URIRef(PDF.set_cgpm_uri(defres))))

# 5) Serialization prefixes
g.serialize(format='turtle', destination=APIPATH + 'prefixes.ttl')
