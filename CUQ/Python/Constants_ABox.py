#
# Constants ABox
#
#
# the module imports the Python script 'settings.py' located in a directory above this directory
# Should this lead to an error when executing the present script (... not found ...) you might need
# to add the location to the PYTHONPATH by typing (in the TERMINAL where you execute the Python script)
# export PYTHONPATH=(path where the package is located) (e.g. /Users/gregordudle/Development/Semantic-SI)
#

from rdflib import *
from CUQ_TBox import SiElements
from datetime import date
from settings import *
import owlrl
import openpyxl
import re
from rdflib.container import Seq

g = Graph()
PDF = SiElements()
BASESTR = str(PDF.BASE_PATH)

# copy over all namespaces from PDF.g to g
for key, val in PDF.g.namespaces():
    g.bind(key, val) 

# Annotations to the ontology (name, creation date, comment)
g.add((URIRef(PDF.namespace_constants), RDF.type, OWL.Ontology))
g.add((URIRef(PDF.namespace_constants), SKOS.prefLabel, Literal("SI Reference Point - Constants", datatype=XSD.string)))
g.add((URIRef(PDF.namespace_constants), RDFS.comment,
           Literal("Ontology, part of the SI reference point, covering the seven underpinning constants of the SI",
                   datatype=XSD.string)))
g.add((URIRef(PDF.namespace_constants), DCTERMS.created, Literal(str(date.today()), datatype=XSD.date)))

# worksheet containing the basic information
cst_wb_obj = openpyxl.load_workbook(XLS_FILES_FOLDER + 'SI_Constants.xlsx')
sheet = cst_wb_obj.active

# load Units graph (to allow identification of URI for a given unit symbol)
units_g = Graph()
units_g.parse(APIPATH + 'si.ttl')
units_g.parse(APIPATH + 'units.ttl')
owlrl.DeductiveClosure(owlrl.OWLRL_Semantics).expand(units_g)

# define auxilary functions


# function to get the power of the unit element
def get_pwr(qty_str: str) -> int:
    try:
        int(qty_str[-1:])
    except ValueError:
        return 1
    else:
        return int(qty_str[-2:])


# function to get the URI of a unit knowing its symbol
def get_uri_for_symbol(sym: str) -> URIRef:
    query = """
        PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>
        PREFIX SI: <http://si-digital-framework.org/SI#>
        PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT ?Unit
        WHERE 
        { 
            ?Unit a SI:MeasurementUnit .
            OPTIONAL{?Unit SI:hasEndValidity ?EndValidity}
            FILTER(!BOUND(?EndValidity))

            ?Unit SI:hasSymbol ?Symbol .
            FILTER (?Symbol='""" + sym + """')
        }"""
    
    qres = units_g.query(query)
    ausgabe = None
    for elment in qres:
        ausgabe = elment['Unit']
    return ausgabe  # error indicates wrong type...


notes_wb_obj = openpyxl.load_workbook(BASESTR + '/_docs/Notes.xlsx')

# A) create dictionary with the note symbol codes (@xxx@) in text
notessym = notes_wb_obj["symbols"]
syms, symtypes, formats = {}, {}, {'html': 6, 'latex': 7, 'json': 8, 'text': 9}
for symrow in range(2, notessym.max_row + 1):
    symtypes.update({notessym.cell(symrow, column=5).value: notessym.cell(symrow, column=4).value})
    for ttype in ['latex']:
        if ttype not in syms.keys():
            syms.update({ttype: {}})
        code = notessym.cell(symrow, column=5).value
        text = notessym.cell(symrow, column=formats[ttype]).value
        syms[ttype].update({code: text})


# B) function to change out symbol references in notes (@xxx@), uses syms and symtypes vars defined above
def formattxt(txt, fmt='html', lvl=0):
    """
    formats a text string with symbols replaced in any of the following formats
    html, latex, json, text
    """
    if fmt not in ['html', 'latex', 'json', 'text']:
        return txt

    matches = re.findall(f"@(.*?)@", txt)
    # order matches by length of substituting string (smallest first) so that replacement of substrings
    # at the top level does not result in multiple replacements at top level (resulting in $$)
    tmatches = {}
    for match in matches:
        strlen = len(syms[fmt][match])
        tmatches.update({match: strlen})
    omatches = dict(sorted(tmatches.items(), key=lambda item: item[1]))
    if omatches:
        for grp in omatches:
            if lvl == 0 and fmt == 'latex':
                # add latex delimiters at the top level
                if symtypes[grp] == 'symbol':
                    txt = txt.replace('@' + grp + '@', "$" + syms[fmt][grp] + "$")
                elif symtypes[grp] == 'equation':
                    txt = txt.replace('@' + grp + '@', "$$" + syms[fmt][grp] + "$$")
            else:
                txt = txt.replace('@' + grp + '@', syms[fmt][grp])
        txt = txt.replace('\n', ' ')
        # check for text still containing symbols based on symbols being parts of other symbols
        if '@' in txt:
            lvl = lvl + 1
            txt = formattxt(txt, fmt, lvl)
    return txt


# crawl through the rows of the XLS file
for row in range(2, sheet.max_row + 1):
    identifier = sheet.cell(row, column=1).value
    hidden_label = sheet.cell(row, column=2).value
    label_en = sheet.cell(row, column=3).value
    label_fr = sheet.cell(row, column=4).value
    unit_type = sheet.cell(row, column=5).value
    value_str = sheet.cell(row, column=6).value
    value = sheet.cell(row, column=7).value
    unit = sheet.cell(row, column=8).value
    unit_str = sheet.cell(row, column=9).value
    symbol = sheet.cell(row, column=10).value
    updated = sheet.cell(row, column=11).value
    defining_res = sheet.cell(row, column=12).value
    element = PDF.set_constant_uri(identifier)

    if "@" in symbol:
        symbol = formattxt(symbol, 'latex')

    g.add((element, RDF.type, PDF.Constant))
    g.add((element, PDF.hasValueAsString, Literal(value_str, datatype=XSD.string)))
    g.add((element, PDF.hasUnitAsString, Literal(unit_str, datatype=XSD.string)))
    if unit_type == "xsd:integer":
        g.add((element, PDF.hasDatatype,  XSD.integer))
        g.add((element, PDF.hasValue, Literal(value, datatype=XSD.integer, normalize=False)))
    elif unit_type == "xsd:double":
        # changing the output datatype to xsd:float as RDFLib has a known bug that loses precision in value
        # https://github.com/RDFLib/rdflib/issues/1852
        g.add((element, PDF.hasDatatype, XSD.float))
        g.add((element, PDF.hasValue, Literal(value, datatype=XSD.float, normalize=False)))
    g.add((element, PDF.hasSymbol, Literal(symbol, datatype=XSD.string)))
    g.add((element, PDF.hasUpdatedDate, Literal(updated, datatype=XSD.date)))
    g.add((element, SKOS.prefLabel, Literal(label_en, lang="en")))
    g.add((element, SKOS.prefLabel, Literal(label_fr, lang="fr")))
    g.add((element, SKOS.hiddenLabel, Literal(hidden_label, datatype=XSD.string)))
    g.add((element, PDF.hasDefiningResolution, PDF.set_cgpm_uri(defining_res)))

    piece_list = []
    pieces = unit.split(".")
    for piece in pieces:
        pwr = get_pwr(piece)
        if pwr != 1:
            piece = piece[:-2]
        blankNodeID = BNode()
        URI_unit = get_uri_for_symbol(piece)
        g.add((blankNodeID, PDF.hasUnit, URI_unit))
        g.add((blankNodeID, PDF.hasUnitPwr, Literal(pwr)))
        piece_list.append(blankNodeID)

    seq_uri = Seq(g, BNode(), piece_list).uri
    g.add((element, PDF.hasUnitElement, seq_uri))

g.serialize(format='turtle', destination=APIPATH + 'constants.ttl')
