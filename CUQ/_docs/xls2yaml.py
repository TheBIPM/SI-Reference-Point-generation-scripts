"""
Transform excel input files into yaml
"""

import openpyxl
from ruamel.yaml import YAML


def sheet2yaml(sheet, output_name, write=True):
    """ Assume first row is list of fields then each row is a record
    """
    output = []
    # This returns something like 'A1:G25'
    dim = sheet.calculate_dimension()
    lr_cell = dim.split(':')[1]
    last_col = ord(lr_cell[0]) - 64
    last_row = int(lr_cell[1:])
    first_row = 1
    # Special cases - pre processing
    if output_name == "def_collectors":
        last_col = 5
    if output_name == "si_units_special_names":
        last_row = 23
        last_col = 10
    if output_name == "non_si_units":
        first_row = 6
    if output_name == "quantities":
        # Dump "same as"
        last_col = 4
    # Get fields list by parsing first row
    field_list = []
    for col in range(1, last_col + 1):
        field_label = sheet.cell(first_row, col).value
        # Avoid special char (poor design by me !)
        if field_label == '#':
            field_label = 'Num'
        if output_name in ["notes_en", "notes_fr"] and field_label == "note":
            field_label = "note_" + output_name.split('_')[1]
        if output_name == "si_constants" and field_label == "value_nist":
            field_label = "value"
        field_list.append(field_label)
    for row in range(first_row + 1, last_row + 1):
        buf = {}
        for i, f in enumerate(field_list):
            buf[f] = sheet.cell(row, i + 1).value
        output.append(buf)
    # Special cases - post processing
    if output_name == "def_collectors":
        # need to append list of definitions
        for row in range(2, last_row + 1):
            output[row - 2]['definitions'] = []
            for col in range(6, 11):
                defid = sheet.cell(row, col).value
                if defid is not None:
                    output[row - 2]['definitions'].append(defid)
    yaml = YAML()
    if write:
        with open(output_name + '.yaml', 'w') as fp:
            yaml.dump(output, fp)
    return output


units_wb_obj = openpyxl.load_workbook('Units_Prefixes.xlsx')
sheet2yaml(units_wb_obj['Prefixes'], 'prefixes')
sheet2yaml(units_wb_obj['BaseUnitsDefs'], 'base_units_defs')
sheet2yaml(units_wb_obj['DefCollectors'], 'def_collectors')
sheet2yaml(units_wb_obj['SIUnitsSpecialNames'], 'si_units_special_names')
sheet2yaml(units_wb_obj['NonSIUnits'], 'non_si_units')

notes_wb_obj = openpyxl.load_workbook('Notes.xlsx')
# Merge notes
notes = sheet2yaml(notes_wb_obj['en'], 'notes_en', write=False)
notes_fr = sheet2yaml(notes_wb_obj['fr'], 'notes_fr', write=False)
for i, n in enumerate(notes):
    notes[i]['note_fr'] = notes_fr[i]['note_fr']
yaml = YAML()
with open('notes.yaml', 'w') as fp:
    yaml.dump(notes, fp)

symbols = sheet2yaml(notes_wb_obj['symbols'], 'symbols', write=False)
symbols_dict = {}
for item in symbols:
    symbols_dict[item['code']] = {
        'name_en': item['name_en'],
        'name_fr': '',
        'type': item['type'],
        'latex': item['latex'],
        'text': item['text'],
        'html': '',
        'json': '',
    }
with open('symbols.yaml', 'w') as fp:
    yaml.dump(symbols_dict, fp)

quantities_wb_obj = openpyxl.load_workbook('quantities.xlsx')
sheet2yaml(quantities_wb_obj['Sheet1'], 'quantities')

missing_def = openpyxl.load_workbook('missing_definitions_cipm.xlsx')
sheet2yaml(missing_def['Feuil1'], 'missing_definitions_cipm')

si_constants = openpyxl.load_workbook('SI_Constants.xlsx')
sheet2yaml(si_constants['Sheet1'], 'si_constants')
